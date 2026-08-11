"""Chapter 4 pipeline — Step 5 of 5: assemble result tables into a formatted
Excel workbook. Writes /mnt/user-data/outputs/Chapter4_results.xlsx.
"""
import pandas as pd
import json
import warnings

warnings.filterwarnings('ignore')

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = 'results/'
schema   = json.load(open('results/schema.json'))
axes_res = json.load(open(OUT + 'out_axes.json'))
XLS      = OUT + 'Chapter4_results.xlsx'

tabs = {
    '01_models':              pd.read_csv(OUT + 'tab_models.csv'),
    '02_metadata_univariate': pd.read_csv(OUT + 'tab_univariate_metadata.csv'),
    '03_supersense_univariate': pd.read_csv(OUT + 'tab_univariate_supersense.csv'),
    '04_LIWC_univariate':     pd.read_csv(OUT + 'tab_univariate_LIWC.csv'),
    '05_foregrounding_univar': pd.read_csv(OUT + 'tab_univariate_foregrounding.csv'),
    '06_semantic_indices':    pd.read_csv(OUT + 'tab_semantic_indices.csv'),
    '07_PCA_variance':        pd.read_csv(OUT + 'tab_pca.csv'),
    '08_PCA_PC1_loadings':    pd.read_csv(OUT + 'tab_pca_pc1_loadings.csv'),
    '09_prestige_test':       pd.read_csv(OUT + 'tab_prestige_test.csv'),
    '10_genre_profiles':      pd.read_csv(OUT + 'tab_genre_profiles.csv'),
    '11_errors_by_genre':     pd.read_csv(OUT + 'tab_errors_by_genre.csv'),
    '12_boundary_nonfic_fic': pd.read_csv(OUT + 'tab_boundary_nonfic_as_fic.csv'),
    '13_boundary_fic_nonfic': pd.read_csv(OUT + 'tab_boundary_fic_as_nonfic.csv'),
    '14_length_confound':     pd.read_csv(OUT + 'tab_length_confound.csv'),
}

readme = pd.DataFrame({
    'Item': [
        'Corpus', 'Corpus size', 'Categories', 'Publication years',
        'Label column',
        'Feature layer 1 — metadata', 'Feature layer 2 — supersense',
        'Feature layer 3 — LIWC',    'Feature layer 4 — foregrounding',
        'Data repair', 'Normalization', 'Validation',
        'Result — fictionality axis', 'Result — prestige axis',
        'Result — foregrounding axis', 'Result — axis orthogonality',
        'Result — reception control', 'Result — boundary residuals',
        'Generated',
    ],
    'Detail': [
        'CONLIT (Piper) — English-language contemporary prose',
        f'{schema["n_works"]:,} works ({schema["n_fic"]:,} FIC / {schema["n_non"]:,} NON)',
        '12 genres: 8 fiction (BS, NYT, PW, MID, ROM, MY, SF, YA), 4 nonfiction (BIO, MEM, HIST, MIX)',
        '2001–2021',
        'Category (FIC / NON) plus 12-way Genre',
        '11 columns: total_characters, protagonist_concentration, avg_sentence_length, avg_word_length, tuldava_score, event_count, speed_avg, circuitousness, speed_min, volume, BAYES_ranking',
        '41 WordNet supersenses (26 noun + 15 verb), normalized per 1,000 tokens',
        '117 LIWC-22 features (WC dropped; correlates with token_count at r = 0.996)',
        '8 unigram-derived proxies: lex_rarity, types_used_rate, hapax_rate, rare_tail_share, work_entropy, KL_from_corpus, KL_from_own_category, composite foregrounding_index',
        f'5 metadata columns carried a decimal-separator artifact (rows inflated 1000x). Repaired: {schema["artifact_repair_log"]}. Corrects circuitousness direction from higher-in-NON to higher-in-FIC (d = +1.12).',
        'Supersenses: raw counts / token_count * 1000. Foregrounding proxies: relative frequencies. LIWC: as delivered (percentages).',
        '10-fold stratified cross-validation. All P(fiction) values are out-of-fold.',
        f'LIWC-based logistic classifier reaches AUC = {axes_res["AUC_fictionality_LIWC"]:.4f} on FIC vs NON.',
        f'LIWC-based logistic classifier trained on PW vs {{ROM, MY, SF, YA, BS}} reaches AUC = {axes_res["AUC_prestige_PW_vs_genre_LIWC"]:.4f}.',
        'Composite of lex_rarity, types_used_rate, hapax_rate, rare_tail_share. Higher = more distinctive vocabulary. Passes the within-fiction validation gate: prizewinners > genre fiction, d = +0.79.',
        f'cosine(fictionality, prestige) on LIWC = {axes_res["cosine_fictionality_prestige_LIWC"]:+.3f}; angle = {axes_res["angle_deg"]:.1f} deg; bootstrap 95% CI [{axes_res["cosine_bootstrap_95CI"][0]:+.2f}, {axes_res["cosine_bootstrap_95CI"][1]:+.2f}]. Two axes are statistically independent.',
        'BAYES_ranking alone: AUC = 0.508 (chance). Reception-only accuracy 0.7028 = majority-class baseline. Dropping BAYES from the full model changes AUC by 0.0003.',
        'Errors concentrate in memoir (10.9%) and mixed nonfiction (6.7%). Within fiction, prizewinners are the most misread (2.7%). Mystery and romance: 0 errors.',
        pd.Timestamp.today().strftime('%Y-%m-%d'),
    ],
})

with pd.ExcelWriter(XLS, engine='openpyxl') as xw:
    readme.to_excel(xw, sheet_name='README', index=False)
    for name, tab in tabs.items():
        tab.to_excel(xw, sheet_name=name, index=False)

# ---------- formatting ----------
wb        = load_workbook(XLS)
hdr_fill  = PatternFill('solid', fgColor='1F4E79')
hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
body_font = Font(name='Arial', size=10)
thin      = Side(style='thin', color='D9D9D9')

for ws in wb.worksheets:
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body_font
            c.border = Border(bottom=thin)
            if isinstance(c.value, float):
                c.number_format = '0.000'
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 60)
    ws.freeze_panes = 'A2'

ws = wb['README']
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 110
for row in ws.iter_rows(min_row=2):
    row[1].alignment = Alignment(wrap_text=True, vertical='top')
    row[0].font      = Font(name='Arial', size=10, bold=True)

wb.save(XLS)
print(f'wrote {XLS}')
print(f'sheets: {len(tabs) + 1}')
